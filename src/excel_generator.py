"""Excel file generation with embedded images."""

from pathlib import Path
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
import tempfile
import os


class ExcelGenerator:
    """Generate Excel files with practice menus and images."""

    def __init__(self):
        """Initialize the Excel generator."""
        self.workbook = Workbook()
        self.temp_files: list[str] = []

    def create_practice_sheet(self, practice_plan: dict, step_images: list[Image.Image]) -> None:
        """Create a practice sheet with images and descriptions.

        Args:
            practice_plan: Practice plan data from LLM.
            step_images: List of PIL Images for each step.
        """
        ws = self.workbook.active
        ws.title = "練習メニュー"

        # Styles
        title_font = Font(name="Meiryo", size=16, bold=True)
        header_font = Font(name="Meiryo", size=12, bold=True)
        normal_font = Font(name="Meiryo", size=10)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(name="Meiryo", size=12, bold=True, color="FFFFFF")

        # Set column widths
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 15

        current_row = 1

        # Title
        ws.merge_cells(f"A{current_row}:D{current_row}")
        title_cell = ws.cell(row=current_row, column=1, value=practice_plan.get("title", "練習メニュー"))
        title_cell.font = title_font
        title_cell.alignment = center_align
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # Description
        ws.merge_cells(f"A{current_row}:D{current_row}")
        desc_cell = ws.cell(row=current_row, column=1, value=practice_plan.get("description", ""))
        desc_cell.font = normal_font
        desc_cell.alignment = left_align
        ws.row_dimensions[current_row].height = 40
        current_row += 2

        # Header row
        headers = ["#", "図解", "説明", "時間"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font_white
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = header_fill
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Steps
        steps = practice_plan.get("steps", [])
        for i, step in enumerate(steps):
            step_start_row = current_row

            # Step number
            ws.cell(row=current_row, column=1, value=step.get("step_number", i + 1))
            ws.cell(row=current_row, column=1).font = header_font
            ws.cell(row=current_row, column=1).alignment = center_align
            ws.cell(row=current_row, column=1).border = thin_border

            # Image
            if i < len(step_images):
                step_data = step_images[i]

                if isinstance(step_data, dict) and "ground" in step_data:
                    # Separate images mode - ground and players are separate
                    self._add_separate_images(ws, current_row, i, step_data)
                else:
                    # Legacy mode - single combined image
                    img = step_data
                    temp_path = self._save_temp_image(img, f"step_{i}.png")
                    xl_img = XLImage(temp_path)

                    # Resize image to fit cell
                    max_width = 350
                    max_height = 200
                    aspect_ratio = img.width / img.height

                    if img.width > max_width:
                        xl_img.width = max_width
                        xl_img.height = int(max_width / aspect_ratio)
                    if xl_img.height > max_height:
                        xl_img.height = max_height
                        xl_img.width = int(max_height * aspect_ratio)

                    # Add image to cell
                    img_cell = f"B{current_row}"
                    ws.add_image(xl_img, img_cell)

            ws.cell(row=current_row, column=2).border = thin_border

            # Description
            step_desc = f"【{step.get('name', '')}】\n\n{step.get('description', '')}"
            ws.cell(row=current_row, column=3, value=step_desc)
            ws.cell(row=current_row, column=3).font = normal_font
            ws.cell(row=current_row, column=3).alignment = left_align
            ws.cell(row=current_row, column=3).border = thin_border

            # Duration
            duration = step.get("duration_minutes", 5)
            ws.cell(row=current_row, column=4, value=f"{duration}分")
            ws.cell(row=current_row, column=4).font = normal_font
            ws.cell(row=current_row, column=4).alignment = center_align
            ws.cell(row=current_row, column=4).border = thin_border

            # Set row height for image
            ws.row_dimensions[current_row].height = 160

            current_row += 1

        # Key points section
        current_row += 1
        ws.merge_cells(f"A{current_row}:D{current_row}")
        kp_header = ws.cell(row=current_row, column=1, value="📌 重要ポイント")
        kp_header.font = header_font
        kp_header.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        key_points = practice_plan.get("key_points", [])
        for point in key_points:
            ws.merge_cells(f"A{current_row}:D{current_row}")
            point_cell = ws.cell(row=current_row, column=1, value=f"• {point}")
            point_cell.font = normal_font
            point_cell.alignment = left_align
            current_row += 1

        # Arrow legend section
        current_row += 1
        ws.merge_cells(f"A{current_row}:D{current_row}")
        legend_header = ws.cell(row=current_row, column=1, value="🔰 図の見方")
        legend_header.font = header_font
        legend_header.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Blue arrow legend
        ws.merge_cells(f"A{current_row}:D{current_row}")
        blue_legend = ws.cell(row=current_row, column=1, value="→ 青色の矢印（実線）: プレイヤーの動き")
        blue_legend.font = Font(name="Meiryo", size=10, color="0000FF")
        blue_legend.alignment = left_align
        current_row += 1

        # Orange arrow legend
        ws.merge_cells(f"A{current_row}:D{current_row}")
        orange_legend = ws.cell(row=current_row, column=1, value="⇢ オレンジ色の矢印（破線）: ボールの動き")
        orange_legend.font = Font(name="Meiryo", size=10, color="FF8C00")
        orange_legend.alignment = left_align
        current_row += 1

    def _add_separate_images(self, ws, row: int, step_idx: int, step_data: dict) -> None:
        """Add ground and player images separately to allow individual movement.

        Args:
            ws: Worksheet to add images to.
            row: Row number to add images at (1-indexed).
            step_idx: Step index for unique temp file names.
            step_data: Dictionary with 'ground', 'players', and 'ground_size'.
        """
        ground_img = step_data["ground"]
        players = step_data["players"]
        ground_width, ground_height = step_data["ground_size"]

        # Calculate scaling to fit in cell
        max_width = 350
        max_height = 200
        aspect_ratio = ground_width / ground_height

        if ground_width > max_width:
            display_width = max_width
            display_height = int(max_width / aspect_ratio)
        else:
            display_width = ground_width
            display_height = ground_height

        if display_height > max_height:
            display_height = max_height
            display_width = int(max_height * aspect_ratio)

        scale = display_width / ground_width

        # Add ground image anchored to cell B{row}
        ground_temp = self._save_temp_image(ground_img, f"step_{step_idx}_ground.png")
        xl_ground = XLImage(ground_temp)
        xl_ground.width = display_width
        xl_ground.height = display_height

        # Use OneCellAnchor - anchor to column B (index 1), row (0-indexed)
        ground_marker = AnchorMarker(col=1, colOff=0, row=row - 1, rowOff=0)
        ground_anchor = OneCellAnchor(
            _from=ground_marker,
            ext=XDRPositiveSize2D(pixels_to_EMU(display_width), pixels_to_EMU(display_height))
        )
        xl_ground.anchor = ground_anchor
        ws.add_image(xl_ground)

        # Add each player as a separate image
        for p_idx, player in enumerate(players):
            player_img = player["image"]
            player_x = player["x"]  # Center position in original coords
            player_y = player["y"]

            # Save player image
            player_temp = self._save_temp_image(
                player_img, f"step_{step_idx}_player_{p_idx}.png"
            )
            xl_player = XLImage(player_temp)

            # Scale player image
            player_display_width = int(player_img.width * scale)
            player_display_height = int(player_img.height * scale)
            xl_player.width = player_display_width
            xl_player.height = player_display_height

            # Calculate offset from cell B{row} (center player on the position)
            offset_x = int(player_x * scale) - player_display_width // 2
            offset_y = int(player_y * scale) - player_display_height // 2

            # Use OneCellAnchor with offset from cell B{row}
            player_marker = AnchorMarker(
                col=1,
                colOff=pixels_to_EMU(offset_x),
                row=row - 1,
                rowOff=pixels_to_EMU(offset_y)
            )
            player_anchor = OneCellAnchor(
                _from=player_marker,
                ext=XDRPositiveSize2D(pixels_to_EMU(player_display_width), pixels_to_EMU(player_display_height))
            )
            xl_player.anchor = player_anchor
            ws.add_image(xl_player)

        # Add each arrow as a separate image
        arrows = step_data.get("arrows", [])
        for a_idx, arrow in enumerate(arrows):
            arrow_img = arrow["image"]
            arrow_x = arrow["x"]  # Center position in original coords
            arrow_y = arrow["y"]

            # Save arrow image (with transparency)
            arrow_temp = self._save_temp_image(
                arrow_img, f"step_{step_idx}_arrow_{a_idx}.png", preserve_alpha=True
            )
            xl_arrow = XLImage(arrow_temp)

            # Scale arrow image
            arrow_display_width = int(arrow_img.width * scale)
            arrow_display_height = int(arrow_img.height * scale)
            xl_arrow.width = arrow_display_width
            xl_arrow.height = arrow_display_height

            # Calculate offset from cell B{row} (center arrow on the position)
            offset_x = int(arrow_x * scale) - arrow_display_width // 2
            offset_y = int(arrow_y * scale) - arrow_display_height // 2

            # Use OneCellAnchor with offset from cell B{row}
            arrow_marker = AnchorMarker(
                col=1,
                colOff=pixels_to_EMU(offset_x),
                row=row - 1,
                rowOff=pixels_to_EMU(offset_y)
            )
            arrow_anchor = OneCellAnchor(
                _from=arrow_marker,
                ext=XDRPositiveSize2D(pixels_to_EMU(arrow_display_width), pixels_to_EMU(arrow_display_height))
            )
            xl_arrow.anchor = arrow_anchor
            ws.add_image(xl_arrow)

    def _save_temp_image(self, img: Image.Image, filename: str, preserve_alpha: bool = False) -> str:
        """Save image to a temporary file.

        Args:
            img: PIL Image to save.
            filename: Filename for the temp file.
            preserve_alpha: If True, preserve transparency (RGBA).

        Returns:
            Path to the temporary file.
        """
        temp_dir = tempfile.gettempdir()
        temp_path = os.path.join(temp_dir, filename)
        if preserve_alpha:
            img.save(temp_path, "PNG")
        else:
            img.convert("RGB").save(temp_path, "PNG")
        self.temp_files.append(temp_path)
        return temp_path

    def save(self, output_path: str | Path) -> Path:
        """Save the workbook to file.

        Args:
            output_path: Path to save the Excel file.

        Returns:
            Path to the saved file.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(output_path)
        return output_path

    def cleanup(self):
        """Clean up temporary files."""
        for temp_file in self.temp_files:
            try:
                os.remove(temp_file)
            except OSError:
                pass
        self.temp_files.clear()

    def __del__(self):
        """Destructor to clean up temp files."""
        self.cleanup()
